import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op
import os


window = tk.Tk()
window.title("Simple Ordering System")
window.configure(bg="lightblue")


FILE_NAME = "Roces_Database.xlsx"

if not os.path.exists(FILE_NAME):
    wb = op.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"])
    wb.save(FILE_NAME)

selected_id = None


title = tk.Label(
    window,
    text="Simple Ordering System",
    font=("Times New Roman", 14, "bold"),
    bg="lightblue"
)
title.grid(row=0, column=0, columnspan=6, pady=10)


genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=6, padx=10, pady=10)


cname_entry = tk.Entry(genframe, font=("Poppins", 12))
cname_entry.grid(row=0, column=0, padx=10, pady=5)

cname_label = tk.Label(
    genframe,
    text="Customer Name",
    font=("Poppins", 10, "italic"),
    bg="lightblue"
)
cname_label.grid(row=1, column=0)


product_entry = tk.Entry(genframe, font=("Poppins", 12))
product_entry.grid(row=0, column=1, padx=10, pady=5)

product_label = tk.Label(
    genframe,
    text="Product",
    font=("Poppins", 10, "italic"),
    bg="lightblue"
)
product_label.grid(row=1, column=1)


qty_entry = tk.Entry(genframe, font=("Poppins", 12))
qty_entry.grid(row=2, column=0, padx=10, pady=5)

qty_label = tk.Label(
    genframe,
    text="Quantity",
    font=("Poppins", 10, "italic"),
    bg="lightblue"
)
qty_label.grid(row=3, column=0)


price_entry = tk.Entry(genframe, font=("Poppins", 12))
price_entry.grid(row=2, column=1, padx=10, pady=5)

price_label = tk.Label(
    genframe,
    text="Price",
    font=("Poppins", 10, "italic"),
    bg="lightblue"
)
price_label.grid(row=3, column=1)


def generate_id():
    wb = op.load_workbook(FILE_NAME)
    ws = wb.active

    if ws.max_row == 1:
        return 1

    return ws.cell(ws.max_row, 1).value + 1


def clear_entries():
    cname_entry.delete(0, tk.END)
    product_entry.delete(0, tk.END)
    qty_entry.delete(0, tk.END)
    price_entry.delete(0, tk.END)


def validate_inputs():

    if cname_entry.get().strip() == "":
        messagebox.showerror("Error", "Customer Name is required!")
        return False

    if product_entry.get().strip() == "":
        messagebox.showerror("Error", "Product is required!")
        return False

    qty = qty_entry.get().strip()

    if not qty.isdigit():
        messagebox.showerror("Error", "Quantity must be a whole number!")
        return False

    if int(qty) <= 0:
        messagebox.showerror("Error", "Quantity must be greater than 0!")
        return False

    price = price_entry.get().strip()

    if price.count(".") > 1:
        messagebox.showerror("Error", "Invalid Price!")
        return False

    if not price.replace(".", "").isdigit():
        messagebox.showerror("Error", "Price must be numeric!")
        return False

    if float(price) <= 0:
        messagebox.showerror("Error", "Price must be greater than 0!")
        return False

    return True


def load_data():
    for item in table.get_children():
        table.delete(item)

    wb = op.load_workbook(FILE_NAME)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=row)


def submit_record():

    if not validate_inputs():
        return

    order_id = generate_id()
    cname = cname_entry.get()
    product = product_entry.get()
    qty = int(qty_entry.get())
    price = float(price_entry.get())
    total = qty * price

    wb = op.load_workbook(FILE_NAME)
    ws = wb.active

    ws.append([order_id, cname, product, qty, price, total])

    wb.save(FILE_NAME)

    load_data()
    clear_entries()

    messagebox.showinfo("Success", "Record Added Successfully!")


def select_record(event):

    global selected_id

    selected = table.focus()

    if not selected:
        return

    values = table.item(selected, "values")

    selected_id = int(values[0])

    clear_entries()

    cname_entry.insert(0, values[1])
    product_entry.insert(0, values[2])
    qty_entry.insert(0, values[3])
    price_entry.insert(0, values[4])


def update_record():

    global selected_id

    if selected_id is None:
        messagebox.showerror("Error", "Select a record first!")
        return

    if not validate_inputs():
        return

    if not messagebox.askyesno("Confirm", "Update selected record?"):
        return

    wb = op.load_workbook(FILE_NAME)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):

        if row[0].value == selected_id:

            qty = int(qty_entry.get())
            price = float(price_entry.get())
            total = qty * price

            row[1].value = cname_entry.get()
            row[2].value = product_entry.get()
            row[3].value = qty
            row[4].value = price
            row[5].value = total

            break

    wb.save(FILE_NAME)

    load_data()
    clear_entries()

    selected_id = None

    messagebox.showinfo("Success", "Record Updated Successfully!")


def delete_record():

    global selected_id

    if selected_id is None:
        messagebox.showerror("Error", "Select a record first!")
        return

    if not messagebox.askyesno("Confirm", "Delete selected record?"):
        return

    wb = op.load_workbook(FILE_NAME)
    ws = wb.active

    for row_num in range(2, ws.max_row + 1):

        if ws.cell(row_num, 1).value == selected_id:
            ws.delete_rows(row_num)
            break

    wb.save(FILE_NAME)

    load_data()
    clear_entries()

    selected_id = None

    messagebox.showinfo("Success", "Record Deleted Successfully!")


submit_btn = tk.Button(
    window,
    text="Submit",
    bg="lightpink",
    font=("Poppins", 12, "bold"),
    command=submit_record
)
submit_btn.grid(row=2, column=1, pady=10)

update_btn = tk.Button(
    window,
    text="Update",
    bg="lightgreen",
    font=("Poppins", 12, "bold"),
    command=update_record
)
update_btn.grid(row=2, column=2)

delete_btn = tk.Button(
    window,
    text="Delete",
    bg="red",
    fg="white",
    font=("Poppins", 12, "bold"),
    command=delete_record
)
delete_btn.grid(row=2, column=3)


table = ttk.Treeview(
    window,
    columns=("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"),
    show="headings",
    height=10
)

for heading in ("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"):
    table.heading(heading, text=heading)

table.grid(row=3, column=0, columnspan=6, padx=10, pady=10)

table.bind("<<TreeviewSelect>>", select_record)


load_data()


window.mainloop()
